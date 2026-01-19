"""Local operational script: Excel -> JSON -> Render API.

Architecture notes (by design):
- This script runs on a Windows machine WITH Excel files.
- The backend NEVER reads .xlsx.
- This script reads the .xlsx and POSTs JSON to the backend.

Supported Excel formats:

1) Normalized format (any sheet)
    Columns (case-insensitive): symbol, name, market, price, price_type, ts_price, source

2) Barchart "Quotes" sheet (as in your screenshot)
    Columns (case-insensitive): Quotes (symbol), Name, Last (price), Timestamp

3) Historical sheets (new)
        Two sheets (defaults):
            - CashHistorical
            - 3MHistorical
        These are Barchart time series sheets and differ from "Quotes".
        Expected columns (case-insensitive):
            - Date + Symbol + Close (preferred)
            - (Open/High/Low/Volume may exist and are ignored)
        Notes:
            - Often row 1 contains a title like "Time Series" and the real column headers
                are on row 2; use --header-row 2 if needed (the script also auto-detects).

Example usage (PowerShell):
  $env:API_BASE_URL = "https://<your-render-app>.onrender.com/api"
  $env:INGEST_TOKEN = "<same token configured on backend>"
  python .\scripts\ingest_lme_from_excel.py --xlsx "C:\lme_market_api\data\market.xlsx"

If your sheet has different headers, pass --sheet and/or --header-row.
"""

from __future__ import annotations

import argparse
import os
from datetime import date, datetime, time, timezone
from typing import Any, Dict, Iterable, List, Optional, Sequence


def _require_pkg(name: str) -> None:
    raise SystemExit(
        f"Missing dependency '{name}'. Install with: pip install {name}"
    )


_ALLOWED_SYMBOLS = {"P3Y00", "P4Y00", "Q7Y00", "^USDBRL"}
_SYMBOL_TO_LIVE_PRICE_TYPE = {
    "P3Y00": "live",
    "P4Y00": "live",
    "Q7Y00": "official",
    "^USDBRL": "live",
}
_SYMBOL_TO_NAME = {
    "P3Y00": "Aluminium Hg Cash",
    "P4Y00": "Aluminium Hg 3M",
    "Q7Y00": "Aluminium Hg Official",
    "^USDBRL": "U.S. Dollar/Brazilian Real",
}


def _norm_header(v: Any) -> str:
    s = "" if v is None else str(v)
    s = s.strip().lower()
    s = s.replace(" ", "_").replace("-", "_")
    while "__" in s:
        s = s.replace("__", "_")
    return s


def _to_utc_iso(ts_val: Any) -> str:
    if isinstance(ts_val, datetime):
        ts = ts_val
        if ts.tzinfo is None:
            ts = ts.replace(tzinfo=timezone.utc)
        return ts.astimezone(timezone.utc).isoformat().replace("+00:00", "Z")

    if isinstance(ts_val, date) and not isinstance(ts_val, datetime):
        ts = datetime.combine(ts_val, time(0, 0), tzinfo=timezone.utc)
        return ts.isoformat().replace("+00:00", "Z")

    s = "" if ts_val is None else str(ts_val).strip()
    if not s:
        raise ValueError("empty timestamp")

    for fmt in (
        "%Y-%m-%dT%H:%M:%S%z",
        "%Y-%m-%dT%H:%M:%S",
        "%Y-%m-%d %H:%M:%S",
        "%Y-%m-%d",
        "%m/%d/%Y",
        "%d/%m/%Y",
        "%m/%d/%y",
        "%d/%m/%y",
    ):
        try:
            dt = datetime.strptime(s, fmt)
            if dt.tzinfo is None:
                dt = dt.replace(tzinfo=timezone.utc)
            return dt.astimezone(timezone.utc).isoformat().replace("+00:00", "Z")
        except ValueError:
            continue

    if s.endswith("Z"):
        try:
            dt = datetime.fromisoformat(s.replace("Z", "+00:00"))
            return dt.astimezone(timezone.utc).isoformat().replace("+00:00", "Z")
        except Exception:
            pass

    raise ValueError(f"unrecognized timestamp format: {s}")


def _load_rows(
    xlsx_path: str,
    sheet_name: str | None,
    header_row: int,
) -> List[Dict[str, Any]]:
    try:
        import openpyxl
    except Exception:
        _require_pkg("openpyxl")

    wb = openpyxl.load_workbook(xlsx_path, data_only=True)
    ws = wb[sheet_name] if sheet_name else wb.worksheets[0]

    # header_row is 1-based
    header_cells = [c.value for c in ws[header_row]]
    headers = [_norm_header(h) for h in header_cells]

    def _idx_any(candidates: List[str]) -> Optional[int]:
        for c in candidates:
            key = _norm_header(c)
            if key in headers:
                return headers.index(key)
        return None

    # Normalized format
    col_symbol = _idx_any(["symbol"])
    col_name = _idx_any(["name"])
    col_market = _idx_any(["market"])
    col_price = _idx_any(["price"])
    col_price_type = _idx_any(["price_type"])
    col_ts_price = _idx_any(["ts_price"])
    col_source = _idx_any(["source"])

    is_normalized = all(
        c is not None
        for c in [
            col_symbol,
            col_name,
            col_market,
            col_price,
            col_price_type,
            col_ts_price,
            col_source,
        ]
    )

    # Barchart Quotes format
    bc_symbol = _idx_any(["quotes", "quote", "symbol"])
    bc_name = _idx_any(["name"])
    bc_last = _idx_any(["last", "last_price"])
    bc_ts = _idx_any(["timestamp", "time", "as_of", "asof"])
    is_barchart = all(c is not None for c in [bc_symbol, bc_name, bc_last, bc_ts])

    if not is_normalized and not is_barchart:
        raise SystemExit(
            "Unsupported sheet layout. Expected either normalized columns "
            "(symbol,name,market,price,price_type,ts_price,source) or Barchart Quotes "
            "(Quotes/Name/Last/Timestamp). Found: "
            + ", ".join([h for h in headers if h])
        )

    out: List[Dict[str, Any]] = []
    for r in ws.iter_rows(min_row=header_row + 1, values_only=True):
        if not any(v is not None and str(v).strip() for v in r):
            continue

        if is_normalized:
            symbol = str(r[col_symbol]).strip()
            if symbol not in _ALLOWED_SYMBOLS:
                continue

            out.append(
                {
                    "symbol": symbol,
                    "name": str(r[col_name]).strip(),
                    "market": str(r[col_market]).strip(),
                    "price": float(r[col_price]),
                    "price_type": str(r[col_price_type]).strip(),
                    "ts_price": _to_utc_iso(r[col_ts_price]),
                    "source": str(r[col_source]).strip(),
                }
            )
            continue

        # Barchart Quotes
        symbol = str(r[bc_symbol]).strip()
        if symbol not in _ALLOWED_SYMBOLS:
            continue

        market = "FX" if symbol.startswith("^") else "LME"

        out.append(
            {
                "symbol": symbol,
                "name": str(r[bc_name]).strip(),
                "market": market,
                "price": float(r[bc_last]),
                "price_type": _SYMBOL_TO_LIVE_PRICE_TYPE[symbol],
                "ts_price": _to_utc_iso(r[bc_ts]),
                "source": "barchart_excel",
            }
        )

    return out


def _resolve_sheet_name(sheet_names: Sequence[str], desired: str) -> str | None:
    desired_key = str(desired or "").strip().casefold()
    if not desired_key:
        return None
    for s in sheet_names:
        if str(s).casefold() == desired_key:
            return s
    return None


def _load_historical_sheet(
    xlsx_path: str,
    sheet_name: str,
    header_row: int,
    *,
    source: str,
) -> List[Dict[str, Any]]:
    try:
        import openpyxl
    except Exception:
        _require_pkg("openpyxl")

    wb = openpyxl.load_workbook(xlsx_path, data_only=True)
    resolved = _resolve_sheet_name(wb.sheetnames, sheet_name)
    if not resolved:
        return []

    ws = wb[resolved]

    def _read_headers(row: int) -> List[str]:
        header_cells = [c.value for c in ws[row]]
        return [_norm_header(h) for h in header_cells]

    def _idx_any(candidates: Iterable[str]) -> Optional[int]:
        for c in candidates:
            key = _norm_header(c)
            if key in headers:
                return headers.index(key)
        return None

    # Some Barchart time series sheets have a title row ("Time Series") then headers.
    candidate_header_rows = [header_row]
    if header_row == 1:
        candidate_header_rows.append(2)
    candidate_header_rows.append(header_row + 1)

    parsed_header_row: int | None = None
    col_date = col_symbol = col_close = None
    headers: List[str] = []
    for hr in candidate_header_rows:
        if hr < 1:
            continue
        headers = _read_headers(hr)
        col_date = _idx_any(["date", "data", "day", "timestamp", "ts", "as_of", "asof"])
        col_symbol = _idx_any(["symbol", "quotes", "quote"])
        # Prefer Close for MTM / daily series.
        col_close = _idx_any(["close", "settle", "settlement", "last", "price", "value"])
        if col_date is not None and col_symbol is not None and col_close is not None:
            parsed_header_row = hr
            break

    if parsed_header_row is None or col_date is None or col_symbol is None or col_close is None:
        raise SystemExit(
            f"Unsupported historical sheet layout in '{resolved}'. Expected Date + Symbol + Close columns. Found: "
            + ", ".join([h for h in headers if h])
        )

    out: List[Dict[str, Any]] = []
    for r in ws.iter_rows(min_row=parsed_header_row + 1, values_only=True):
        if not any(v is not None and str(v).strip() for v in r):
            continue

        ts_val = r[col_date]
        symbol_val = r[col_symbol]
        close_val = r[col_close]
        if ts_val is None or symbol_val is None or close_val is None:
            continue

        symbol = str(symbol_val).strip()
        if symbol not in _ALLOWED_SYMBOLS:
            continue

        try:
            price_f = float(close_val)
        except Exception:
            continue

        # Historical time series are used for chart + MTM close (D-1).
        # Keep them distinct from intraday "live" quotes.
        price_type = "official" if symbol == "Q7Y00" else "close"

        market = "FX" if symbol.startswith("^") else "LME"

        out.append(
            {
                "symbol": symbol,
                "name": _SYMBOL_TO_NAME[symbol],
                "market": market,
                "price": price_f,
                "price_type": price_type,
                "ts_price": _to_utc_iso(ts_val),
                "source": source,
            }
        )

    return out


def main() -> int:
    p = argparse.ArgumentParser()
    p.add_argument("--xlsx", required=True, help="Path to market.xlsx")
    p.add_argument(
        "--sheet",
        default="Quotes",
        help="Live quotes sheet name (default: Quotes; matches Barchart export)",
    )
    p.add_argument(
        "--cash-historical-sheet",
        default="CashHistorical",
        help="Cash historical sheet name (default: CashHistorical)",
    )
    p.add_argument(
        "--three-month-historical-sheet",
        default="3MHistorical",
        help="3M historical sheet name (default: 3MHistorical)",
    )
    p.add_argument(
        "--usdbrl-historical-sheet",
        default="USDBRL",
        help="USDBRL historical sheet name (default: USDBRL)",
    )
    p.add_argument(
        "--header-row",
        type=int,
        default=1,
        help="1-based row number containing column headers",
    )
    p.add_argument(
        "--no-live",
        action="store_true",
        help="Skip live quotes ingest (Quotes sheet)",
    )
    p.add_argument(
        "--no-history",
        action="store_true",
        help="Skip historical ingest (cashhistorical / 3mhistorical sheets)",
    )
    p.add_argument(
        "--api-base-url",
        default=os.getenv("API_BASE_URL", "").rstrip("/"),
        help="Base API URL, e.g. https://.../api (or set API_BASE_URL)",
    )
    p.add_argument(
        "--token",
        default=os.getenv("INGEST_TOKEN", ""),
        help="Ingest token (or set INGEST_TOKEN)",
    )

    args = p.parse_args()

    if not args.api_base_url:
        raise SystemExit("API base URL required via --api-base-url or API_BASE_URL")
    if not args.token:
        raise SystemExit("Ingest token required via --token or INGEST_TOKEN")

    rows: List[Dict[str, Any]] = []
    if not args.no_live:
        rows.extend(_load_rows(args.xlsx, args.sheet, args.header_row))
    if not args.no_history:
        rows.extend(
            _load_historical_sheet(
                args.xlsx,
                args.cash_historical_sheet,
                args.header_row,
                source="barchart_excel_cashhistorical",
            )
        )
        rows.extend(
            _load_historical_sheet(
                args.xlsx,
                args.three_month_historical_sheet,
                args.header_row,
                source="barchart_excel_3mhistorical",
            )
        )
        rows.extend(
            _load_historical_sheet(
                args.xlsx,
                args.usdbrl_historical_sheet,
                args.header_row,
                source="barchart_excel_usdbrl",
            )
        )

    if not rows:
        raise SystemExit(
            "No rows found to ingest (check sheet names and header row)."
        )

    try:
        import requests
    except Exception:
        _require_pkg("requests")

    url = f"{args.api_base_url}/ingest/lme/price"
    headers = {"Authorization": f"Bearer {args.token}"}

    ok = 0
    fail = 0
    for row in rows:
        resp = requests.post(url, json=row, headers=headers, timeout=20)
        if 200 <= resp.status_code < 300:
            ok += 1
        else:
            fail += 1
            print("FAILED", resp.status_code, row.get("symbol"), resp.text[:300])

    print(f"done ok={ok} fail={fail} total={len(rows)}")
    return 0 if fail == 0 else 2


if __name__ == "__main__":
    raise SystemExit(main())
